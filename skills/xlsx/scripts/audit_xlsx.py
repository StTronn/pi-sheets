#!/usr/bin/env python3
"""audit_xlsx — workbook formula audit (broken refs + cached formula errors).

Library use:
    from audit_xlsx import audit
    report = audit("workbook.xlsx")

CLI:
    python scripts/audit_xlsx.py <workbook.xlsx>
"""
from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Union

from openpyxl import load_workbook


ERROR_PREFIXES = (
    "#NULL!", "#DIV/0!", "#VALUE!", "#REF!", "#NAME?",
    "#NUM!", "#N/A", "#GETTING_DATA", "#SPILL!", "#CALC!", "#ERROR!",
)


class AuditError(Exception):
    def __init__(self, message: str, *, code: str = "error"):
        super().__init__(message)
        self.code = code


def _is_error_value(value: object) -> bool:
    return isinstance(value, str) and value.startswith(ERROR_PREFIXES)


def audit(wb_path: Union[str, Path]) -> dict:
    """Audit workbook formulas. Returns the same dict the CLI prints.

    Raises AuditError on missing file, bad extension, or open failure.
    """
    wb_path = Path(wb_path).expanduser().resolve()
    if not wb_path.exists():
        raise AuditError(f"workbook not found: {wb_path}", code="not_found")
    if wb_path.suffix.lower() != ".xlsx":
        raise AuditError(
            f"expected .xlsx workbook, got: {wb_path.suffix or '<no extension>'}",
            code="bad_extension",
        )

    try:
        wb_formula = load_workbook(wb_path, data_only=False)
        wb_cached = load_workbook(wb_path, data_only=True)
    except Exception as exc:
        raise AuditError(f"failed to open workbook: {exc}", code="open_failed")

    sheet_reports: list[dict] = []
    formula_cells = 0
    formula_error_cells: list[dict] = []
    ref_error_cells: list[dict] = []

    for ws_formula, ws_cached in zip(wb_formula.worksheets, wb_cached.worksheets):
        sheet_formula_count = 0
        sheet_error_count = 0
        sheet_ref_count = 0

        for row in ws_formula.iter_rows():
            for cell in row:
                value = cell.value
                if not (isinstance(value, str) and value.startswith("=")):
                    continue
                sheet_formula_count += 1
                formula_cells += 1

                cached_value = ws_cached[cell.coordinate].value
                if _is_error_value(cached_value):
                    sheet_error_count += 1
                    formula_error_cells.append({
                        "sheet": ws_formula.title,
                        "cell": cell.coordinate,
                        "formula": value,
                        "cached_value": cached_value,
                    })

                if "#REF!" in value.upper():
                    sheet_ref_count += 1
                    ref_error_cells.append({
                        "sheet": ws_formula.title,
                        "cell": cell.coordinate,
                        "formula": value,
                    })

        sheet_reports.append({
            "sheet": ws_formula.title,
            "max_row": ws_formula.max_row,
            "max_column": ws_formula.max_column,
            "formula_cells": sheet_formula_count,
            "formula_error_cells": sheet_error_count,
            "ref_error_formulas": sheet_ref_count,
        })

    return {
        "ok": not formula_error_cells and not ref_error_cells,
        "path": str(wb_path),
        "sheet_count": len(wb_formula.worksheets),
        "formula_cells": formula_cells,
        "formula_error_count": len(formula_error_cells),
        "ref_error_count": len(ref_error_cells),
        "sheets": sheet_reports,
        "formula_errors": formula_error_cells,
        "ref_errors": ref_error_cells,
    }


def _fail(message: str, *, code: str = "error") -> None:
    print(json.dumps({"ok": False, "code": code, "message": message}))
    raise SystemExit(1)


def main() -> None:
    if len(sys.argv) != 2:
        _fail("usage: audit_xlsx.py <workbook.xlsx>", code="usage")
    try:
        result = audit(sys.argv[1])
    except AuditError as exc:
        _fail(str(exc), code=exc.code)
    print(json.dumps(result))


if __name__ == "__main__":
    main()
