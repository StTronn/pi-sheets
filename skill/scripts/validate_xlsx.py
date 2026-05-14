#!/usr/bin/env python3
"""validate_xlsx — single-pass audit + formualizer recalc in one call.

Library use:
    from validate_xlsx import validate
    report = validate("workbook.xlsx")

CLI:
    python scripts/validate_xlsx.py <workbook.xlsx> [--timeout-sec N]

Output: dict (or JSON to stdout):
    {
      ok, path,
      audit:  { formula_cells, formula_error_count, ref_error_count, formula_errors, ref_errors },
      recalc: { formula_cells, formula_error_count, errors },
      elapsed_ms
    }
Exit: 0 = clean, 1 = any errors or fatal error.
"""
from __future__ import annotations

import argparse
import json
import time
from pathlib import Path
from typing import Union

from openpyxl import load_workbook

from recalc_xlsx import RecalcError, recalc

_ERROR_PREFIXES = (
    "#NULL!", "#DIV/0!", "#VALUE!", "#REF!", "#NAME?",
    "#NUM!", "#N/A", "#GETTING_DATA", "#SPILL!", "#CALC!", "#ERROR!",
)


class ValidateError(Exception):
    def __init__(self, message: str, *, code: str = "error"):
        super().__init__(message)
        self.code = code


def _is_error_value(value: object) -> bool:
    return isinstance(value, str) and value.startswith(_ERROR_PREFIXES)


def _audit(wb_path: Path) -> dict:
    """Walk wb (formula + cached views) once, return audit dict."""
    wb_formula = load_workbook(wb_path, data_only=False)
    wb_cached = load_workbook(wb_path, data_only=True)

    formula_error_cells: list[dict] = []
    ref_error_cells: list[dict] = []
    formula_cells = 0

    for ws_f, ws_c in zip(wb_formula.worksheets, wb_cached.worksheets):
        for row in ws_f.iter_rows():
            for cell in row:
                v = cell.value
                if not (isinstance(v, str) and v.startswith("=")):
                    continue
                formula_cells += 1
                cached = ws_c[cell.coordinate].value
                if _is_error_value(cached):
                    formula_error_cells.append({
                        "sheet": ws_f.title, "cell": cell.coordinate,
                        "formula": v, "cached_value": cached,
                    })
                if "#REF!" in v.upper():
                    ref_error_cells.append({
                        "sheet": ws_f.title, "cell": cell.coordinate, "formula": v,
                    })
    return {
        "formula_cells": formula_cells,
        "formula_error_count": len(formula_error_cells),
        "ref_error_count": len(ref_error_cells),
        "formula_errors": formula_error_cells,
        "ref_errors": ref_error_cells,
    }


def validate(wb_path: Union[str, Path]) -> dict:
    """Run audit + formualizer recalc. Returns combined report dict.

    Raises ValidateError on missing file / bad extension / open failure,
    or re-raises RecalcError from the recalc step.
    """
    wb_path = Path(wb_path).expanduser().resolve()
    if not wb_path.exists():
        raise ValidateError(f"workbook not found: {wb_path}", code="not_found")
    if wb_path.suffix.lower() != ".xlsx":
        raise ValidateError(
            f"expected .xlsx, got: {wb_path.suffix or '<no ext>'}",
            code="bad_extension",
        )

    started = time.time()

    try:
        audit_report = _audit(wb_path)
    except Exception as exc:
        raise ValidateError(f"failed to read workbook: {exc}", code="open_failed")

    recalc_payload = recalc(wb_path)  # raises RecalcError on fatal

    ok = (
        audit_report["formula_error_count"] == 0
        and audit_report["ref_error_count"] == 0
        and recalc_payload.get("ok", False)
    )
    return {
        "ok": ok,
        "path": str(wb_path),
        "audit": audit_report,
        "recalc": {
            "formula_cells": recalc_payload.get("formula_cells", 0),
            "formula_error_count": recalc_payload.get("formula_error_count", 0),
            "errors": recalc_payload.get("errors", []),
        },
        "elapsed_ms": int((time.time() - started) * 1000),
    }


def _fail(message: str, *, code: str = "error") -> None:
    print(json.dumps({"ok": False, "code": code, "message": message}))
    raise SystemExit(1)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("path")
    parser.add_argument("--timeout-sec", type=int, default=45)  # parity-only
    args = parser.parse_args()

    try:
        result = validate(args.path)
    except (ValidateError, RecalcError) as exc:
        _fail(str(exc), code=exc.code)

    print(json.dumps(result))
    raise SystemExit(0 if result["ok"] else 1)


if __name__ == "__main__":
    main()
