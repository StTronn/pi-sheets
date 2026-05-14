#!/usr/bin/env python3
"""
recalc_xlsx.py — formula recalc + error detection via the formualizer
(Rust) engine. In-process; no Node/LibreOffice subprocess.

Usage:
    python scripts/recalc_xlsx.py <workbook.xlsx>
        [--timeout-sec N]                 # accepted for CLI parity (unused)
        [--write-evaluated PATH]          # value-only xlsx with native types

Output (stdout): JSON
    { ok, path, formula_cells, formula_error_count,
      errors: [{sheet, coord, formula, value}],
      elapsed_ms, evaluated_path?, evaluated_path_error? }
Exit: 0 = clean, 1 = formula errors or fatal error.
"""
from __future__ import annotations

import argparse
import json
import os
import tempfile
import time
from pathlib import Path

import formualizer as fz
from openpyxl import load_workbook


# Workaround for formualizer 0.5.8: unary `+` on string operands returns #VALUE!
# instead of pass-through. Excel/LibreOffice both pass through. The idiom
# `=+SheetA!Cell` is very common in finmodels (Lotus-1-2-3 carry-over) and
# breaks downstream when the source cell is a string label (e.g. "2014F").
# Conservative scope: rewrite only the leading-`=+` case, leave nested `+`
# inside expressions alone (would need a real parser to handle safely).
# Tracking: PSU3D0/formualizer issue (to be filed).
def _should_strip_leading_plus(formula: str) -> bool:
    return (
        formula.startswith("=+")
        and len(formula) > 2
        and formula[2] not in "+-="  # skip "=++", "=+-", "=+="
    )


# Excel error code returned when evaluate_cell yields {'type':'Error','kind':K}.
# Mapping verified against formualizer 0.5.8 — kinds observed: Div, Name, Ref, Na,
# Null, Num, Value, Spill, Calc, Error (catch-all incl. #GETTING_DATA).
_KIND_TO_CODE = {
    "Div": "#DIV/0!",
    "Name": "#NAME?",
    "Ref": "#REF!",
    "Na": "#N/A",
    "Null": "#NULL!",
    "Num": "#NUM!",
    "Value": "#VALUE!",
    "Spill": "#SPILL!",
    "Calc": "#CALC!",
    "Error": "#ERROR!",
    "Cancelled": "#ERROR!",
    "Circ": "#REF!",
    "Nimpl": "#NAME?",
}


def _fail(message: str, *, code: str = "error") -> None:
    print(json.dumps({"ok": False, "code": code, "message": message}))
    raise SystemExit(1)


def _excel_code(ev: object) -> str | None:
    """Return Excel error string if ev is an error sentinel, else None."""
    if isinstance(ev, dict) and ev.get("type") == "Error":
        return _KIND_TO_CODE.get(ev.get("kind", ""), "#ERROR!")
    return None


def _coerce_native(v: object) -> object:
    """formualizer returns float/int/bool/str/dict already — pass through."""
    return v


class RecalcError(Exception):
    """Recalc-step failure with a stable error code for callers."""
    def __init__(self, message: str, *, code: str = "error"):
        super().__init__(message)
        self.code = code


def recalc(wb_path: Path, write_evaluated: Path | None = None) -> dict:
    """Run formualizer recalc on wb_path. Returns the same dict main() prints.

    Raises RecalcError on fatal failure (file missing, parse failure, etc.).
    """
    wb_path = Path(wb_path).expanduser().resolve()
    if not wb_path.exists():
        raise RecalcError(f"workbook not found: {wb_path}", code="not_found")
    if wb_path.suffix.lower() != ".xlsx":
        raise RecalcError(f"expected .xlsx, got: {wb_path.suffix or '<no ext>'}",
                          code="bad_extension")

    started = time.time()

    # 1. Discover formula cells via openpyxl (need formula text + coordinates
    #    for the error report and value-rewrite step). One pass.
    try:
        wb_op = load_workbook(wb_path, data_only=False)
    except Exception as exc:
        raise RecalcError(f"failed to read workbook: {exc}", code="open_failed")

    # formulas[sheet_name] = list of (row1, col1, coord, formula_text)
    formulas: dict[str, list[tuple[int, int, str, str]]] = {}
    formula_cells = 0
    for ws in wb_op.worksheets:
        bucket: list[tuple[int, int, str, str]] = []
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith("="):
                    bucket.append((cell.row, cell.column, cell.coordinate, v))
        if bucket:
            formulas[ws.title] = bucket
            formula_cells += len(bucket)

    # 1.5. Workaround for formualizer unary-`+` bug. If any formula starts with
    #      `=+<non-op>`, write a temp copy with those rewritten to `=<...>` and
    #      point formualizer at the temp. wb_op is left untouched so a later
    #      `write_evaluated` pass still sees user-original formulas.
    unary_plus_rewrites = 0
    fz_load_path: Path = wb_path
    tmp_path: str | None = None
    needs_rewrite = any(
        _should_strip_leading_plus(ftxt)
        for cells in formulas.values() for (_, _, _, ftxt) in cells
    )
    if needs_rewrite:
        wb_for_fz = load_workbook(wb_path, data_only=False)
        for ws in wb_for_fz.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if isinstance(v, str) and _should_strip_leading_plus(v):
                        cell.value = "=" + v[2:]
                        unary_plus_rewrites += 1
        fd, tmp_path = tempfile.mkstemp(suffix=".xlsx", prefix="recalc_")
        os.close(fd)
        wb_for_fz.save(tmp_path)
        fz_load_path = Path(tmp_path)

    # 2. Load + evaluate via formualizer
    try:
        try:
            fwb = fz.Workbook.load_path(str(fz_load_path))
            fwb.evaluate_all()
        except Exception as exc:
            raise RecalcError(f"formualizer evaluation failed: {exc}", code="eval_failed")
    finally:
        if tmp_path:
            try:
                os.unlink(tmp_path)
            except OSError:
                pass

    fz_sheets = set(fwb.sheet_names)

    # 3. Walk formula cells, collect errors + computed values
    errors: list[dict] = []
    # values[sheet] = {coord: native_value}; skip sheets with >5000 formulas
    # to keep stdout/file size sane.
    values: dict[str, dict[str, object]] = {}

    for sheet_name, cells in formulas.items():
        if sheet_name not in fz_sheets:
            # Sheet present in xlsx but not seen by formualizer — flag every formula
            for (_, _, coord, ftxt) in cells:
                errors.append({"sheet": sheet_name, "coord": coord,
                               "formula": ftxt, "value": "#REF!"})
            continue

        collect = len(cells) <= 5000
        if collect:
            values[sheet_name] = {}

        for (r, c, coord, ftxt) in cells:
            try:
                ev = fwb.evaluate_cell(sheet_name, r, c)
            except Exception as exc:
                errors.append({"sheet": sheet_name, "coord": coord,
                               "formula": ftxt, "value": f"#ERROR!({exc})"})
                continue

            err_code = _excel_code(ev)
            if err_code is not None:
                errors.append({"sheet": sheet_name, "coord": coord,
                               "formula": ftxt, "value": err_code})
            elif collect and ev is not None:
                values[sheet_name][coord] = _coerce_native(ev)

    # 4. Optional value-only xlsx with native types
    evaluated_path: str | None = None
    evaluated_path_error: str | None = None
    if write_evaluated:
        try:
            out = Path(write_evaluated).expanduser().resolve()
            out.parent.mkdir(parents=True, exist_ok=True)
            for ws in wb_op.worksheets:
                sheet_vals = values.get(ws.title)
                if not sheet_vals:
                    continue
                for row in ws.iter_rows():
                    for cell in row:
                        v = cell.value
                        if not (isinstance(v, str) and v.startswith("=")):
                            continue
                        computed = sheet_vals.get(cell.coordinate)
                        if computed is not None:
                            cell.value = computed
            wb_op.save(out)
            evaluated_path = str(out)
        except Exception as exc:
            evaluated_path_error = str(exc)

    result: dict = {
        "ok": len(errors) == 0,
        "path": str(wb_path),
        "formula_cells": formula_cells,
        "formula_error_count": len(errors),
        "errors": errors,
        "elapsed_ms": int((time.time() - started) * 1000),
    }
    if unary_plus_rewrites > 0:
        result["unary_plus_workaround_applied"] = unary_plus_rewrites
    if evaluated_path is not None:
        result["evaluated_path"] = evaluated_path
    if evaluated_path_error is not None:
        result["evaluated_path_error"] = evaluated_path_error
    return result


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("path")
    ap.add_argument("--timeout-sec", type=int, default=45)  # parity-only
    ap.add_argument("--write-evaluated", metavar="PATH",
                    help="Write a value-only xlsx (formulas replaced by computed values) to PATH")
    args = ap.parse_args()

    try:
        result = recalc(Path(args.path),
                        write_evaluated=Path(args.write_evaluated) if args.write_evaluated else None)
    except RecalcError as exc:
        _fail(str(exc), code=exc.code)

    print(json.dumps(result))
    raise SystemExit(0 if result["ok"] else 1)


if __name__ == "__main__":
    main()
