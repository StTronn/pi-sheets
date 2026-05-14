#!/usr/bin/env python3
"""
inspect_xlsx — fast workbook structure summary for agent orientation.

Library use:
    from inspect_xlsx import inspect
    info = inspect("workbook.xlsx")

CLI:
    python scripts/inspect_xlsx.py <workbook.xlsx>

Output: dict (or JSON to stdout for the CLI):
    {
      ok, path, sheet_count,
      sheets: [{ name, dims, rows, cols, formula_cells, data_cells }],
      summary: "SheetA(A1:P45,38f) SheetB(A1:AH80,142f) ..."
    }
"""
from __future__ import annotations

import json
import sys
from pathlib import Path
from typing import Union

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


class InspectError(Exception):
    """Inspect-step failure with a stable error code for CLI translation."""
    def __init__(self, message: str, *, code: str = "error"):
        super().__init__(message)
        self.code = code


def inspect(wb_path: Union[str, Path]) -> dict:
    """Inspect workbook structure. Returns the same dict the CLI prints.

    Raises InspectError on missing file, bad extension, or open failure.
    """
    wb_path = Path(wb_path).expanduser().resolve()
    if not wb_path.exists():
        raise InspectError(f"workbook not found: {wb_path}", code="not_found")
    if wb_path.suffix.lower() != ".xlsx":
        raise InspectError(
            f"expected .xlsx, got: {wb_path.suffix or '<no ext>'}",
            code="bad_extension",
        )

    try:
        wb = load_workbook(wb_path, data_only=False, read_only=True)
    except Exception as exc:
        raise InspectError(f"failed to open workbook: {exc}", code="open_failed")

    sheet_infos: list[dict] = []
    summary_parts: list[str] = []

    for ws in wb.worksheets:
        formula_cells = 0
        data_cells = 0
        min_r = min_c = float("inf")
        max_r = max_c = 0
        first_row_data: list | None = None
        first_row_num: int | None = None

        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                r, c = cell.row, cell.column
                if r < min_r: min_r = r
                if r > max_r: max_r = r
                if c < min_c: min_c = c
                if c > max_c: max_c = c
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_cells += 1
                else:
                    data_cells += 1
            if first_row_data is None:
                row_vals = [cell.value for cell in row if cell.value is not None]
                if row_vals:
                    first_row_data = [cell.value for cell in row][:30]
                    first_row_num = next((cell.row for cell in row if hasattr(cell, "row")), None)

        if max_r == 0:
            summary_parts.append(f"{ws.title}(empty)")
            sheet_infos.append({"name": ws.title, "dims": "empty", "rows": 0, "cols": 0, "formula_cells": 0, "data_cells": 0})
            continue

        top_left = f"{get_column_letter(int(min_c))}{int(min_r)}"
        bot_right = f"{get_column_letter(int(max_c))}{int(max_r)}"
        dims = f"{top_left}:{bot_right}"
        rows = int(max_r - min_r + 1)
        cols = int(max_c - min_c + 1)

        summary_parts.append(f"{ws.title}({dims},{formula_cells}f)")
        sheet_infos.append({
            "name": ws.title,
            "dims": dims,
            "rows": rows,
            "cols": cols,
            "formula_cells": formula_cells,
            "data_cells": data_cells,
            "first_row": first_row_num,
            "first_row_values": first_row_data,
        })

    wb.close()

    return {
        "ok": True,
        "path": str(wb_path),
        "sheet_count": len(sheet_infos),
        "sheets": sheet_infos,
        "summary": " ".join(summary_parts),
    }


def _fail(message: str, *, code: str = "error") -> None:
    print(json.dumps({"ok": False, "code": code, "message": message}))
    raise SystemExit(1)


def main() -> None:
    if len(sys.argv) != 2:
        _fail("usage: inspect_xlsx.py <workbook.xlsx>", code="usage")
    try:
        result = inspect(sys.argv[1])
    except InspectError as exc:
        _fail(str(exc), code=exc.code)
    print(json.dumps(result))


if __name__ == "__main__":
    main()
