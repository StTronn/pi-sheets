#!/usr/bin/env python3
"""extend_formula — copy a formula across a target range with A1-style shifts.

Wraps `openpyxl.formula.translate.Translator` to behave like Excel fill-right
or fill-down.

Library use:
    from extend_formula import extend_formula
    r = extend_formula("workbook.xlsx", source_cell="J16",
                       target_range="K16:O16", sheet=None, overwrite=False)

CLI:
    python3 scripts/extend_formula.py <workbook.xlsx>
        --source-cell J16 --target-range K16:O16
        [--sheet SHEET] [--overwrite]

Output: dict (or JSON to stdout):
    {
      ok, sheet, source_cell, source_formula,
      written: [{coord, formula}, ...],
      skipped: [{coord, reason}, ...],
      errors:  [{coord, error}, ...]
    }
Exit: 0 on success, 1 on fatal failure.
"""
from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Union

from openpyxl import load_workbook
from openpyxl.formula.translate import Translator
from openpyxl.utils import range_boundaries, get_column_letter


class ExtendFormulaError(Exception):
    def __init__(self, message: str, *, code: str = "error"):
        super().__init__(message)
        self.code = code


def _expand_range(target_range: str) -> list[str]:
    min_col, min_row, max_col, max_row = range_boundaries(target_range)
    coords: list[str] = []
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            coords.append(f"{get_column_letter(c)}{r}")
    return coords


def extend_formula(
    wb_path: Union[str, Path],
    source_cell: str,
    target_range: str,
    sheet: str | None = None,
    overwrite: bool = False,
) -> dict:
    """Copy a formula across a range. Saves wb in place when any cell is written.

    Raises ExtendFormulaError on file / sheet / cell / range / save issues.
    """
    wb_path = Path(wb_path).expanduser().resolve()
    if not wb_path.exists():
        raise ExtendFormulaError(f"workbook not found: {wb_path}", code="not_found")
    if wb_path.suffix.lower() != ".xlsx":
        raise ExtendFormulaError(
            f"expected .xlsx, got: {wb_path.suffix or '<no ext>'}",
            code="bad_extension",
        )

    try:
        wb = load_workbook(wb_path, data_only=False)
    except Exception as exc:
        raise ExtendFormulaError(f"failed to read workbook: {exc}", code="open_failed")

    sheet_name = sheet or wb.active.title
    if sheet_name not in wb.sheetnames:
        raise ExtendFormulaError(
            f"sheet not found: {sheet_name!r}; have {wb.sheetnames!r}",
            code="sheet_not_found",
        )
    ws = wb[sheet_name]

    src = ws[source_cell]
    src_value = src.value
    if not (isinstance(src_value, str) and src_value.startswith("=")):
        raise ExtendFormulaError(
            f"source cell {sheet_name}!{source_cell} does not contain a formula "
            f"(value={src_value!r})",
            code="source_not_formula",
        )

    try:
        target_coords = _expand_range(target_range)
    except Exception as exc:
        raise ExtendFormulaError(
            f"bad target_range {target_range!r}: {exc}", code="bad_range"
        )

    translator = Translator(src_value, origin=source_cell)
    written: list[dict] = []
    skipped: list[dict] = []
    errors: list[dict] = []

    for coord in target_coords:
        if coord == source_cell:
            skipped.append({"coord": coord, "reason": "is source cell"})
            continue
        existing = ws[coord].value
        if existing is not None and not overwrite:
            skipped.append({"coord": coord, "reason": f"non-empty (value={existing!r})"})
            continue
        try:
            shifted = translator.translate_formula(coord)
        except Exception as exc:
            errors.append({"coord": coord, "error": str(exc)})
            continue
        ws[coord] = shifted
        written.append({"coord": coord, "formula": shifted})

    if written or errors:
        try:
            wb.save(wb_path)
        except Exception as exc:
            raise ExtendFormulaError(f"failed to save workbook: {exc}", code="save_failed")

    return {
        "ok": len(errors) == 0,
        "sheet": sheet_name,
        "source_cell": source_cell,
        "source_formula": src_value,
        "written": written,
        "skipped": skipped,
        "errors": errors,
    }


def _fail(msg: str, *, code: str = "error") -> None:
    print(json.dumps({"ok": False, "code": code, "message": msg}))
    raise SystemExit(1)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("path")
    ap.add_argument("--source-cell", required=True, help="e.g. J16")
    ap.add_argument("--target-range", required=True, help="e.g. K16:O16")
    ap.add_argument("--sheet", help="sheet name; default = active sheet")
    ap.add_argument("--overwrite", action="store_true",
                    help="overwrite non-empty target cells (default: skip)")
    args = ap.parse_args()

    try:
        result = extend_formula(
            args.path,
            source_cell=args.source_cell,
            target_range=args.target_range,
            sheet=args.sheet,
            overwrite=args.overwrite,
        )
    except ExtendFormulaError as exc:
        _fail(str(exc), code=exc.code)

    print(json.dumps(result))
    raise SystemExit(0 if result["ok"] else 1)


if __name__ == "__main__":
    main()
